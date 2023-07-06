# This class receives values from the parser (threw stemmer) and it builds a dictionary
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import numpy as np
import math as m
from sklearn.metrics.pairwise import cosine_similarity


def run_indexer():
    #  Get all the songs names and delete the'\n' like always
    songs_names = open('songsRunTimeFetch.txt', 'r')
    songs_names_list = songs_names.readlines()
    songs_names_list_mani = []

    songs_fetch_time_list = []  # list to save the fetch time

    # Remove the runtime from the last index, and save all to a new list
    for song_name in songs_names_list:  # For every song
        song_name = song_name[:-1]  # Delete '\n' from the end of the String
        song_name_list = song_name.split()  # Split the new String to array
        song_fetch_time = song_name_list[-1]  # Fetch time
        songs_fetch_time_list.append(song_fetch_time)  # Add the fetch ime to the list
        temp = song_name_list[:-1]  # Remove the runtime index
        songs_names_list_mani.append(" ".join(temp))

    # The following will help us build a dictionary of words the right way
    # 1. list_of_songs -> will represent where the word was and how many times
    # 2. all_words_dict -> will keep all the word and a list for theirs song:
    # "wordName": dance, [[song1, 2], [song2, 4]....]
    # "wordName": grind, [[song5, 8], [song6, 7]....]
    # set_of_songs = []
    # all_words_dict = {"wordName": "word",
    #                   "songs and appearances": set_of_songs}  # A dictionary to save all the words from all the lyrics files

    all_words_dict = {}  # My dictionary
    tf_df_words_dict = {}  # Dictionary that contains: { "Word name": [The list of songs and appearances = tf,
    #                                                                   number of document the word appears = df}
    #

    # Work break down:
    # 1. Go on every song
    # 2. on every song check every word frequency
    #   2.1 Add to a list the song name and its frequency
    #   2.2 Add the list to the list that saves the word's songs appearances and frequency
    # 3. add all unique words to the dictionary

    # dictionary properties
    # key: song's name
    # value: a list containing list that has two values:
    #   [song name, number of appearances of the word in the song]

    songs_names_list_mani = songs_names_list_mani[0:10]
    list_of_total_words = []

    index = 0
    # For loop for arranging the songs names
    for song in songs_names_list_mani:

        try:
            song = str(index) * 3 + " " + song
            temp_song = open(song + '.txt', 'r')
            temp_song_lyrics = (temp_song.readlines())  # Reads all the lines from the file
            temp_song_lyrics_manipulated = []  # new list that will contain all the sentences without '\n
            for sentence in temp_song_lyrics:
                temp_song_lyrics_manipulated.append(sentence.strip())

            list_of_total_words.append(len(temp_song_lyrics_manipulated))  # Saves the total words of the song

            # For loop for adding the words to the dictionary
            for Cword in temp_song_lyrics_manipulated:

                count = temp_song_lyrics_manipulated.count(Cword)  # The word's frequency
                new_value = [song, count]  # word new key -> word: "dance", value: {[song1, 2], [song2, 3]}
                #                                                                   set of lists
                try:  # 'try' section is because I will get 'None' and then exception if I try to get the value...
                    # ...for a not existing key
                    #  If the 'word' is already in the dictionary
                    word_value = all_words_dict[Cword]  # Get it's key
                    if new_value not in word_value:  # Add it if it's not there
                        word_value.append(new_value)
                except:
                    # If it's a new word
                    new_values_list = [new_value]  # Create a new list of list
                    all_words_dict[Cword] = new_values_list  # Add it as key to the new 'word'

                # new_list = [[song, count]]
                # all_words_dict[word] = new_list
                #
                # while count != 0:
                #     temp_song_lyrics_manipulated.remove(word)
                #     count -= 1
        except:
            print(song, "NOT FOUND")

        index += 1  # Increment the index for building the right string for the next song

    list_of_total_words_np = np.array(list_of_total_words)  # Creating a numpy array to normalize the final matrix

    # print(len(songs_names_list_mani))
    # print(len(all_words_dict))

    # First I will sort the dict so the words will be listed lexicography
    sorted_dict = sorted(all_words_dict.items())
    # print(len(sorted_dict))

    # Start of tf df extraction
    # Build dictionary that consist with: { "Word name": [The list of songs and appearances = tf,
    #     #                                               number of document the word appears = df}
    #  final_words_dict
    for b_word in sorted_dict:  # For all the words in the sorted dictionary
        b_word_list_of_lists = b_word[1]  # Grab the list of song list

        b_df = len(b_word_list_of_lists)
        b_tf = 0  # The number of times the word appear in all the words
        for b_word_List in b_word_list_of_lists:  # For every song in the song list
            b_tf += b_word_List[1]  # Update tf

        tf_df_words_dict[b_word[0]] = [b_tf, b_df]  # insert the dictionary for every word its 'tf' and 'df' values
    #  End of tf fd extractions

    sorted_words = sorted(all_words_dict)  # this list will be the first col in the Excel file

    # Build Excel file
    # setting up thw excel workbook for words frequency
    wb = Workbook()
    work_book_bin = wb.active
    work_book_bin.title = 'Words freq'

    wb2 = Workbook()
    work_book_appearances = wb2.active
    work_book_appearances.title = 'Words appearances'

    wb3 = Workbook()
    work_book_not_normalized = wb3.active
    work_book_not_normalized.title = 'Work book not normalized'

    a_songs_indexes = []
    a_words_indexes = []
    i = 1

    work_book_appearances['A1'] = 'wordName'
    # This is the row of all the songs names
    for col in range(0, len(songs_names_list_mani)):  # For enry index in the list
        new_col = get_column_letter(col + 2)  # Jump to the second col (the first on is empty for the words)
        # make the value of each cell
        work_book_bin[new_col + "1"] = songs_names_list_mani[col]  # Words frequency file
        work_book_appearances[new_col + "1"] = songs_names_list_mani[col]  # Words appearances file
        work_book_not_normalized[new_col + "1"] = songs_names_list_mani[col]  # work_book_not_normalized file
        a_songs_indexes.append(i)
        i += 1

    # ********************************* check for deletion
    i = 1
    # This is the first col in the Excel file - all the words
    # Because I already have the words name columns I can fill their frequency and appearances
    for row in range(0, len(sorted_words)):  # for every word
        # Set the value in each row
        a_current_word = sorted_words[row]
        # work_book_bin["A" + str(row + 2)] = sorted_words[row]  # Words frequency file
        # work_book_appearances["A" + str(row + 2)] = sorted_words[row]  # Words appearances file
        a_words_indexes.append(i)
        i += 1
    # ********************************* check for deletion

    matrix_not_normalized = []
    matrix_of_freq = []
    matrix_of_bin = []

    temp_list = []
    temp_list.append(" ")
    temp_list = temp_list + songs_names_list_mani
    matrix_of_freq.append(temp_list)
    matrix_of_bin.append(temp_list)

    x = 0
    for a_s in sorted_dict:  # For all the words in the dictionary
        x += 1
        a_s_key = a_s[0]  # get the word's key. for example: 'man' or 'tree'
        a_s_value = a_s[1]  # Get the word's list (of lists) - meaning the all the songs its appear and frequency
        a_row_freq = [0] * len(songs_names_list_mani)  # A list that represent the word's freq in every song
        # a_row = [0] * len(songs_names_list_mani)  # A list that represent the word's freq in every song
        for a_s_value_list in a_s_value:  # Now, for all the songs that the word appears
            a_s_song = a_s_value_list[0].split()[1:]  # Get the song name - ditch the 000, 111 ....
            a_s_song = " ".join(a_s_song)  # Connect the song name back together
            a_s_song_freq = a_s_value_list[1]  # Get the frequency
            a_index = songs_names_list_mani.index(a_s_song)  # The index of the song that the word appear
            a_row_freq[a_index] = a_s_song_freq

        # Appearances
        work_book_not_normalized.append([a_s_key] + a_row_freq)  # Save the word's appearances (not normalized)
        matrix_not_normalized.append(a_row_freq)
        a_row_freq_np = np.array(a_row_freq)  # Turn the whole row to a numpy array - easy to divide
        a_row_freq = a_row_freq_np / list_of_total_words  # Divide each row with the 'list_of_total_words' words np array - Normalize
        a_row_freq = [a_s_key] + list(a_row_freq)  # Adds the 'word' to the row, turn the row to a regular list

        matrix_of_freq.append(a_row_freq)  # Adds the row to 'matrix_of_freq' matrix
        work_book_appearances.append(a_row_freq)  # Adds the row to the 'work_book_appearances' Excel file

        # "Yes" or "No" in a given song
        # If the value of a cell is more than 0 -> replace the value with '1' because the word appear in that song
        for cell in range(1, len(a_row_freq)):
            if a_row_freq[cell] > 0:
                a_row_freq[cell] = 1

        # Add the rows to 'matrix_of_bin' and the Excel file
        matrix_of_bin.append(a_row_freq)
        work_book_bin.append(a_row_freq)

    list_of_total_words = ["Total song's words"] + list_of_total_words
    work_book_appearances.append(list_of_total_words)

    songs_fetch_time_list = songs_fetch_time_list[:10]
    songs_fetch_time_list = ["Fetch time"] + songs_fetch_time_list
    work_book_appearances.append(songs_fetch_time_list)

    # tf df Excel file
    wb4 = Workbook()
    work_book_tf_df = wb4.active
    work_book_tf_df.title = 'work book tf df'

    work_book_tf_df['A1'] = "Word name"
    work_book_tf_df['B1'] = "tf"
    work_book_tf_df['C1'] = "df"

    for li in tf_df_words_dict:  # Insert tf df values to 'tf_df' Excel file
        work_book_tf_df.append([li] + tf_df_words_dict[li])

    wb.save('Words in song bin.xlsx')
    wb2.save('Words appearances.xlsx')
    wb3.save('Words not normalized.xlsx')
    wb4.save('work book tf df.xlsx')

    # print("point of debugger")
    # return matrix_of_freq, matrix_of_bin
    return sorted_dict, sorted_words, songs_names_list_mani


def find_cosine(w_matrix_np, user_cosine_vector, songs_names_list_mani_ranker):
    w_matrix_np_T = w_matrix_np.T  # The transpose of the original matrix to go threw all the columns that are now rows
    ranked_songs = {}  # The final dictionary that will contain all the words and their cosine similarity
    index = 0  # Index to find the song's name
    for row in w_matrix_np_T:  # For every word in the transpose matrix -> every row is a vector
        cosine = cosine_similarity([row], [user_cosine_vector])  # Find the angle between the vectors
        if cosine != 0:  # If the angle has a meaning
            ranked_songs[
                songs_names_list_mani_ranker[index]] = cosine  # Add the song name and its cosine to the dictionary
            index += 1  # increment the index for the next song's name
        else:
            # If the cosine is 0 then there is no meaning for adding the song to the dictionary
            index += 1

    ranked_songs = dict(sorted(ranked_songs.items(), key=lambda x: x[1], reverse=True))
    return ranked_songs


def print_dictionary(dictionary):  # This method prints only dictionaries
    if type(dictionary) == dict:  # Only if the DS is Dictionary
        for k in dictionary:  # For all the items
            print(" ".join(k.split()[:-1]))  # Split the string to array, delete the 'lyrics' word,...
            # ...and put the string back together
    else:
        print("Can only print Dictionaries")


def ranker():
    N = 10  # Number of documents in my DB
    # Get the right data from the Indexer
    # 'sorted_words_and_songs_list_ranker' is the dictionary: # "wordName": dance, [[song1, 2], [song2, 4]....]
    #                                                           "wordName": grind, [[song5, 8], [song6, 7]....]
    # 'sorted_words_ranker' is Only the words
    sorted_words_and_songs_list_ranker, sorted_words_ranker, songs_names_list_mani_ranker = run_indexer()

    print()
    userInput = input("Please enter some words, separated by spaces...").lower().split()
    if len(userInput) > 0:
        # userInput = ['baby', 'back']  # The user's string
        userInput2 = list(userInput)  # The user's string

        user_input_length = len(userInput)

        # Get the words and their normalized appearances
        wb = load_workbook('D:/pythonCode/Indexer/Words appearances.xlsx')
        Words_appearances = wb.active

        wb2 = load_workbook('D:/pythonCode/Indexer/work book tf df.xlsx')
        work_book_tf_df = wb2.active

        rows_list = []  # List for words appearances
        rows_list2 = []  # List for 'tf' and 'df'
        for i in range(1, len(sorted_words_ranker)):  # Go through every word
            # This 'if' will stop me if I found all the word the user entered
            # Later I will remove the word that I found from the array of the user input, if the array is empty-
            # all the words found and I can stop the loop
            if len(userInput) == 0:
                break
            str_value_i = get_column_letter(i)
            row_a = Words_appearances[str(i)]  # Changing 'i' to str will help me iterate the Excel lines.
            # row_a is the first Excel line.. then second...
            for first_word_in_row in row_a:  # For every word in row
                v = first_word_in_row.value  # Check the first word - 'man', 'baby'...
                if first_word_in_row.value in userInput:  # if the first word is a word that the user entered
                    row = []  # Array for the whole row
                    for index_value in row_a:  # Add all the line to the array
                        row.append(index_value.value)
                    rows_list.append(row)
                    userInput.remove(first_word_in_row.value)
                else:
                    break

            # Fetching tf and df
            row_b = work_book_tf_df[str(i)]  # Changing 'i' to str will help me iterate the Excel lines.
            for first_word_in_row in row_b:  # For every word in row
                v = first_word_in_row.value  # Check the first word - 'man', 'baby'...
                if first_word_in_row.value in userInput2:  # if the first word is a word that the user entered
                    row2 = []  # Array for the whole row
                    for index_value in row_b:  # Add all the line to the array
                        row2.append(index_value.value)
                    rows_list2.append(row2[1:])
                    userInput2.remove(first_word_in_row.value)
                else:
                    break

        w_matrix_demonstraion = []  # This is a matrix for building the demonstration Excel file
        w_matrix_demonstraion.append([" "] + songs_names_list_mani_ranker)

        w_matrix = []  # This matrix will hold the values for the 'Cosine similarity' check

        row_number = 0
        for final_row in rows_list:  # go through every word in 'row_list'
            first_value_final_row = [final_row[0]]  # Remove the word name from the row
            temp_final_row = list(final_row[1:])  # Get only the df values
            df_tf_row = rows_list2[row_number]
            for final_cell in temp_final_row:  # for every cell in the row
                if final_cell != 0:  # If the cell is meaningful
                    tf = df_tf_row[0]
                    df = df_tf_row[1]
                    x = final_cell * (m.log(N / df))
                    temp_final_row[temp_final_row.index(final_cell)] = x

            w_matrix.append(temp_final_row)  # For cosine similarity check
            final_row = first_value_final_row + temp_final_row  # For the Excel file
            w_matrix_demonstraion.append(final_row)
            row_number += 1

        # Create a Excel document for demonstrating the data
        wb5 = Workbook()
        work_book_ranker = wb5.active
        work_book_ranker.title = 'ranker'
        for row in w_matrix_demonstraion:
            work_book_ranker.append(row)
        wb5.save('ranker.xlsx')

        # Cosine similarity - both data vectors and the user's vector must be 2D
        w_matrix_np = np.array(w_matrix)  # the original matrix with all the wighted values

        # Set up the user vector
        words_left = len(userInput)  # How many words left in the array that I remove words from

        dif = user_input_length - words_left  # The difference between the words that left and the original input size

        if dif > 0:  # If dif is 0 then all the word that the user wrote are not in the word's dictionary
            # ... is the size of the vector
            value = 1 / dif
            user_cosine_vector = []  # the user vector
            while dif > 0:  # Adding values to the vector
                user_cosine_vector.append(value)
                dif -= 1

            ranked_songs = find_cosine(w_matrix_np, user_cosine_vector, songs_names_list_mani_ranker)

            print_dictionary(ranked_songs)
        else:
            print("Didn't found any match")  # Any word that the user wrote is not in the dictionary
    else:
        print("Next time write somthing... :(")


ranker()
